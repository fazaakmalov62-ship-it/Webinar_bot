import telebot
from telebot import types
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from flask import Flask, request

# === Настройки ===
BOT_TOKEN = os.environ.get("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN не задан. Установите переменную окружения BOT_TOKEN в Render или локально.")

ADMIN_ID = int(os.environ.get("ADMIN_ID", "1309971729"))
FILE_NAME = "webinar_registrations.xlsx"
RENDER_APP_NAME = os.environ.get("RENDER_APP_NAME")  # например webinar-bot-1juu

bot = telebot.TeleBot(BOT_TOKEN)

# === Создаём файл Excel, если его нет ===
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Никнейм", "Дата регистрации", "ID TG", "Имя", "Статус"])
    wb.save(FILE_NAME)

# === Flask ===
app = Flask(__name__)

@app.route("/", methods=["GET"])
def index():
    return "Bot server is running", 200

# === Webhook endpoint с логированием ===
@app.route(f"/{BOT_TOKEN}", methods=["POST"])
def webhook():
    json_str = request.get_data().decode("utf-8")
    print("Incoming update:", json_str)
    if not json_str:
        return "no data", 400
    try:
        update = telebot.types.Update.de_json(json_str)
        bot.process_new_updates([update])
    except Exception as e:
        print("Ошибка обработки апдейта:", e)
    return "ok", 200

# === Хэндлеры ===
@bot.message_handler(commands=['start'])
def start(message):
    try:
        print(f"Start command received from {message.from_user.id}")
        user_id = message.from_user.id
        wb = load_workbook(FILE_NAME)
        ws = wb.active

        found = False
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=3).value == user_id:
                found = True
                name = ws.cell(row=i, column=4).value or "—"
                status = ws.cell(row=i, column=5).value or ""
                break

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton("Записаться на вебинар"))
        markup.add(types.KeyboardButton("Обновить данные"))
        markup.add(types.KeyboardButton("Отказаться от вебинара"))

        if found:
            bot.send_message(message.chat.id,
                             f"👋 Вы уже зарегистрированы!\nИмя: {name}\nСтатус: {status}",
                             reply_markup=markup)
        else:
            bot.send_message(message.chat.id,
                             "👋 Привет! Нажмите кнопку ниже, чтобы записаться на вебинар:",
                             reply_markup=markup)
    except Exception as e:
        print("Ошибка в /start:", e)

# === Остальные хэндлеры (обёрнуты в try/except аналогично) ===
@bot.message_handler(func=lambda msg: msg.text == "Записаться на вебинар")
def register_step1(message):
    try:
        user_id = message.from_user.id
        username = message.from_user.username or "—"
        reg_time = datetime.now().strftime("%Y-%m-%d %H:%M")

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        found = False
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=3).value == user_id:
                ws.cell(row=i, column=1, value=username)
                ws.cell(row=i, column=2, value=reg_time)
                found = True
                break

        if not found:
            ws.append([username, reg_time, user_id, "", ""])
            try:
                bot.send_message(ADMIN_ID, f"Новый участник зарегистрировался: @{username} ({user_id})")
            except Exception:
                pass

        wb.save(FILE_NAME)

        bot.send_message(message.chat.id, "✍️ Пожалуйста, введите ваше имя:")
        bot.register_next_step_handler(message, register_step2)
    except Exception as e:
        print("Ошибка в register_step1:", e)

def register_step2(message):
    try:
        name = message.text
        user_id = message.from_user.id

        wb = load_workbook(FILE_NAME)
        ws = wb.active
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=3).value == user_id:
                ws.cell(row=i, column=4, value=name)
                ws.cell(row=i, column=5, value="")  
                break
        wb.save(FILE_NAME)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton("Отказаться от вебинара"))
        markup.add(types.KeyboardButton("Обновить данные"))
        bot.send_message(message.chat.id,
                         "✅ Спасибо за регистрацию! Ваши данные сохранены.",
                         reply_markup=markup)
    except Exception as e:
        print("Ошибка в register_step2:", e)

@bot.message_handler(func=lambda msg: msg.text == "Отказаться от вебинара")
def cancel_registration(message):
    try:
        user_id = message.from_user.id
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=3).value == user_id:
                ws.cell(row=i, column=5, value="Отказался")
                break
        wb.save(FILE_NAME)
        bot.send_message(message.chat.id, "❌ Вы отказались от вебинара. Данные обновлены.")
    except Exception as e:
        print("Ошибка в cancel_registration:", e)

@bot.message_handler(func=lambda msg: msg.text == "Обновить данные")
def update_data(message):
    try:
        register_step1(message)
    except Exception as e:
        print("Ошибка в update_data:", e)

@bot.message_handler(commands=['broadcast'])
def broadcast(message):
    try:
        if message.from_user.id != ADMIN_ID:
            bot.send_message(message.chat.id, "❌ У вас нет доступа.")
            return
        msg = bot.send_message(message.chat.id, "Введите текст рассылки:")
        bot.register_next_step_handler(msg, send_broadcast)
    except Exception as e:
        print("Ошибка в broadcast:", e)

def send_broadcast(message):
    try:
        text = message.text
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        sent = 0
        for i in range(2, ws.max_row + 1):
            user_id = ws.cell(row=i, column=3).value
            status = ws.cell(row=i, column=5).value
            if status != "Отказался":
                try:
                    bot.send_message(user_id, text)
                    sent += 1
                except Exception:
                    continue
        bot.send_message(message.chat.id, f"Рассылка отправлена {sent} пользователям.")
    except Exception as e:
        print("Ошибка в send_broadcast:", e)

# === Установка вебхука ===
if RENDER_APP_NAME:
    WEBHOOK_URL = f"https://{RENDER_APP_NAME}.onrender.com/{BOT_TOKEN}"
    try:
        bot.remove_webhook()
        ok = bot.set_webhook(url=WEBHOOK_URL)
        print("set_webhook result:", ok, "WEBHOOK_URL:", WEBHOOK_URL)
    except Exception as e:
        print("Ошибка при установке вебхука:", e)
else:
    print("RENDER_APP_NAME не задан — вебхук не устанавливается автоматически.")

# === Запуск ===
if __name__ == "__main__":
    if os.environ.get("LOCAL_TEST"):
        print("Запуск в режиме polling (LOCAL_TEST).")
        bot.infinity_polling()
    else:
        port = int(os.environ["PORT"])
        app.run(host="0.0.0.0", port=port, threaded=True)
